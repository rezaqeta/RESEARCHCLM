"""
Fast PDF downloader for research papers from Excel files.
Downloads PDFs one by one and marks them in the Excel file.
"""
import pandas as pd
import requests
from pathlib import Path
import time
from urllib.parse import quote_plus, urlparse
import re
from openpyxl import load_workbook

# Configuration
EXCEL_FILES = [
    'source_main_cb.xlsx',
    'Climate_Behavioral_Change_Papers.xlsx'
]
PDF_DOWNLOAD_DIR = Path('sources/metadata/pdfs')
PDF_DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)

# Headers to mimic browser
HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
    'Accept-Language': 'en-US,en;q=0.5',
    'Connection': 'keep-alive',
}

def clean_filename(title, max_length=100):
    """Clean paper title to create valid filename."""
    # Remove invalid characters
    filename = re.sub(r'[<>:"/\\|?*]', '', title)
    # Replace spaces and limit length
    filename = filename.replace(' ', '_')[:max_length]
    return f"{filename}.pdf"

def search_google_scholar_pdf(paper_title):
    """Search Google Scholar for PDF link."""
    search_urls = [
        f"https://scholar.google.com/scholar?q={quote_plus(paper_title)}+filetype:pdf",
        f"https://www.google.com/search?q={quote_plus(paper_title)}+filetype:pdf",
        f"https://www.google.com/search?q={quote_plus(paper_title)}+pdf",
    ]
    return search_urls

def try_direct_link(link):
    """Try downloading from direct link if available."""
    if pd.isna(link) or not link or link == 'nan':
        return None
    
    try:
        # Check if link is already a PDF
        if link.lower().endswith('.pdf'):
            return link
        
        # Try to find PDF link in common repositories
        if 'doi.org' in link:
            # Try to extract DOI and build Sci-Hub link
            doi_match = re.search(r'10\.\d{4,}/[^\s]+', link)
            if doi_match:
                doi = doi_match.group(0)
                return f"https://sci-hub.se/{doi}"
        
        if 'arxiv.org' in link:
            # Convert to PDF link
            pdf_link = link.replace('/abs/', '/pdf/') + '.pdf'
            return pdf_link
        
    except Exception as e:
        print(f"    Error processing link: {e}")
    
    return None

def download_pdf(url, filename):
    """Download PDF from URL."""
    try:
        response = requests.get(url, headers=HEADERS, timeout=30, allow_redirects=True)
        
        # Check if response is PDF
        content_type = response.headers.get('Content-Type', '')
        if 'application/pdf' not in content_type and not url.endswith('.pdf'):
            return False
        
        if response.status_code == 200 and len(response.content) > 1000:  # At least 1KB
            filepath = PDF_DOWNLOAD_DIR / filename
            with open(filepath, 'wb') as f:
                f.write(response.content)
            print(f"    ✓ Downloaded: {filename} ({len(response.content)/1024:.1f} KB)")
            return True
        else:
            print(f"    ✗ Failed: Status {response.status_code}")
            return False
            
    except Exception as e:
        print(f"    ✗ Error downloading: {e}")
        return False

def update_excel_status(excel_file, row_index, status):
    """Update the PDF Available column in Excel file."""
    try:
        # Load workbook with openpyxl to preserve formatting
        wb = load_workbook(excel_file)
        ws = wb.active
        
        # Find PDF Available column (usually column F, index 6)
        # Add 2 to row_index: +1 for 1-based indexing, +1 for header row
        excel_row = row_index + 2
        ws.cell(row=excel_row, column=6, value=status)
        
        wb.save(excel_file)
        return True
    except Exception as e:
        print(f"    ⚠ Could not update Excel: {e}")
        return False

def process_excel_file(excel_file):
    """Process one Excel file - search and download PDFs."""
    print(f"\n{'='*80}")
    print(f"Processing: {excel_file}")
    print(f"{'='*80}\n")
    
    # Read Excel
    df = pd.read_excel(excel_file)
    
    # Filter rows that need PDFs
    mask = (df['PDF Available'].isna()) | (df['PDF Available'] == 'Not Available')
    rows_to_process = df[mask]
    
    print(f"Total papers: {len(df)}")
    print(f"Need PDFs: {len(rows_to_process)}\n")
    
    downloaded_count = 0
    
    for idx, row in rows_to_process.iterrows():
        paper_title = row['Paper Title']
        link = row.get('Link', None)
        paper_no = row.get('No', idx + 1)
        
        print(f"[{paper_no}/{len(df)}] {paper_title[:70]}...")
        
        # Create filename
        filename = clean_filename(paper_title)
        
        # Check if already downloaded
        if (PDF_DOWNLOAD_DIR / filename).exists():
            print(f"    ✓ Already exists: {filename}")
            update_excel_status(excel_file, idx, "Yes")
            downloaded_count += 1
            continue
        
        # Try 1: Direct link from Excel
        pdf_url = try_direct_link(link)
        if pdf_url:
            print(f"    → Trying direct link...")
            if download_pdf(pdf_url, filename):
                update_excel_status(excel_file, idx, "Yes")
                downloaded_count += 1
                time.sleep(1)  # Rate limiting
                continue
        
        # Try 2: Search URLs (user can manually check these)
        search_urls = search_google_scholar_pdf(paper_title)
        print(f"    → Search URLs generated:")
        for i, url in enumerate(search_urls[:2], 1):
            print(f"      {i}. {url[:100]}...")
        
        # Mark as searched but not downloaded
        update_excel_status(excel_file, idx, "Searched - Manual Check Needed")
        
        # Small delay between requests
        time.sleep(0.5)
    
    print(f"\n{'='*80}")
    print(f"Summary for {excel_file}:")
    print(f"  Downloaded: {downloaded_count}")
    print(f"  Manual check needed: {len(rows_to_process) - downloaded_count}")
    print(f"{'='*80}\n")

def main():
    """Main function to process all Excel files."""
    print("\n" + "="*80)
    print("PDF DOWNLOADER - Starting...")
    print("="*80)
    
    for excel_file in EXCEL_FILES:
        if Path(excel_file).exists():
            process_excel_file(excel_file)
        else:
            print(f"⚠ File not found: {excel_file}")
    
    print("\n" + "="*80)
    print("COMPLETE!")
    print(f"PDFs saved to: {PDF_DOWNLOAD_DIR.absolute()}")
    print("="*80)

if __name__ == "__main__":
    main()

