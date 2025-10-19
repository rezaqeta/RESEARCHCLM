"""
Automated PDF downloader with web scraping capabilities.
Searches Google Scholar and other sources, downloads PDFs automatically.
"""
import pandas as pd
import requests
from pathlib import Path
import time
from urllib.parse import quote_plus, urlparse, urljoin
import re
from openpyxl import load_workbook
from bs4 import BeautifulSoup
import json

# Configuration
EXCEL_FILES = [
    'source_main_cb.xlsx',
    'Climate_Behavioral_Change_Papers.xlsx'
]
PDF_DOWNLOAD_DIR = Path('sources/metadata/pdfs')
PDF_DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)

# Headers to mimic browser
HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
    'Accept-Language': 'en-US,en;q=0.5',
    'Connection': 'keep-alive',
}

def clean_filename(title, paper_no=None, max_length=100):
    """Clean paper title to create valid filename."""
    # Remove invalid characters
    filename = re.sub(r'[<>:"/\\|?*]', '', title)
    # Replace spaces and limit length
    filename = filename.replace(' ', '_')[:max_length]
    if paper_no:
        filename = f"{paper_no:03d}_{filename}"
    return f"{filename}.pdf"

def extract_pdf_links_from_google_scholar(paper_title):
    """Search Google Scholar and extract PDF links."""
    search_url = f"https://scholar.google.com/scholar?q={quote_plus(paper_title)}"
    pdf_links = []
    
    try:
        response = requests.get(search_url, headers=HEADERS, timeout=15)
        if response.status_code == 200:
            soup = BeautifulSoup(response.text, 'html.parser')
            
            # Find all PDF links
            for link in soup.find_all('a'):
                href = link.get('href', '')
                text = link.get_text().lower()
                
                # Check if it's a PDF link
                if 'pdf' in text or href.endswith('.pdf'):
                    if href.startswith('http'):
                        pdf_links.append(href)
                    elif href.startswith('/'):
                        pdf_links.append(urljoin('https://scholar.google.com', href))
            
            # Also check for direct links in the results
            for div in soup.find_all('div', class_='gs_or_ggsm'):
                link = div.find('a')
                if link and link.get('href'):
                    pdf_links.append(link.get('href'))
                    
    except Exception as e:
        print(f"    ⚠ Scholar search error: {e}")
    
    return pdf_links

def search_sci_hub(paper_title, doi=None):
    """Try to find paper on Sci-Hub."""
    sci_hub_domains = [
        'https://sci-hub.se',
        'https://sci-hub.st',
        'https://sci-hub.ru'
    ]
    
    search_term = doi if doi else paper_title
    
    for domain in sci_hub_domains:
        try:
            url = f"{domain}/{quote_plus(search_term)}"
            response = requests.get(url, headers=HEADERS, timeout=10)
            if response.status_code == 200:
                soup = BeautifulSoup(response.content, 'html.parser')
                # Find the PDF embed or download link
                iframe = soup.find('iframe', {'id': 'pdf'})
                if iframe and iframe.get('src'):
                    pdf_url = iframe['src']
                    if pdf_url.startswith('//'):
                        pdf_url = 'https:' + pdf_url
                    elif pdf_url.startswith('/'):
                        pdf_url = domain + pdf_url
                    return pdf_url
        except Exception as e:
            continue
    
    return None

def try_arxiv_link(link):
    """Convert arXiv abstract link to PDF."""
    if 'arxiv.org' in link:
        if '/abs/' in link:
            return link.replace('/abs/', '/pdf/') + '.pdf'
        elif 'arxiv.org/pdf' not in link and not link.endswith('.pdf'):
            # Try to construct PDF link
            arxiv_id = re.search(r'(\d{4}\.\d{4,5})', link)
            if arxiv_id:
                return f"https://arxiv.org/pdf/{arxiv_id.group(1)}.pdf"
    return None

def extract_doi_from_link(link):
    """Extract DOI from various link formats."""
    if not link or pd.isna(link):
        return None
    
    # Common DOI patterns
    doi_patterns = [
        r'10\.\d{4,}/[^\s]+',
        r'doi\.org/(10\.\d{4,}/[^\s]+)'
    ]
    
    for pattern in doi_patterns:
        match = re.search(pattern, str(link))
        if match:
            return match.group(1) if 'doi.org' in pattern else match.group(0)
    
    return None

def download_pdf(url, filename, max_size_mb=50):
    """Download PDF from URL with size limit."""
    try:
        # First, check content type and size
        head_response = requests.head(url, headers=HEADERS, timeout=10, allow_redirects=True)
        content_type = head_response.headers.get('Content-Type', '').lower()
        content_length = int(head_response.headers.get('Content-Length', 0))
        
        # Check if it's a PDF and reasonable size
        if content_length > max_size_mb * 1024 * 1024:
            print(f"    ⚠ File too large: {content_length / 1024 / 1024:.1f} MB")
            return False
        
        # Download the file
        response = requests.get(url, headers=HEADERS, timeout=30, allow_redirects=True, stream=True)
        
        if response.status_code == 200:
            # Check if response is actually PDF
            content_type = response.headers.get('Content-Type', '').lower()
            
            # Save to file
            filepath = PDF_DOWNLOAD_DIR / filename
            with open(filepath, 'wb') as f:
                for chunk in response.iter_content(chunk_size=8192):
                    f.write(chunk)
            
            # Verify it's a PDF by checking magic bytes
            with open(filepath, 'rb') as f:
                magic = f.read(4)
                if magic == b'%PDF':
                    file_size = filepath.stat().st_size
                    print(f"    ✓ Downloaded: {filename} ({file_size/1024:.1f} KB)")
                    return True
                else:
                    # Not a PDF, delete it
                    filepath.unlink()
                    print(f"    ✗ Not a PDF file")
                    return False
        else:
            print(f"    ✗ HTTP {response.status_code}")
            return False
            
    except Exception as e:
        print(f"    ✗ Download error: {str(e)[:50]}")
        return False

def update_excel_status(excel_file, row_index, status):
    """Update the PDF Available column in Excel file."""
    try:
        wb = load_workbook(excel_file)
        ws = wb.active
        excel_row = row_index + 2
        ws.cell(row=excel_row, column=6, value=status)
        wb.save(excel_file)
        return True
    except Exception as e:
        print(f"    ⚠ Excel update failed: {e}")
        return False

def find_and_download_pdf(paper_title, link, paper_no):
    """Try multiple methods to find and download PDF."""
    filename = clean_filename(paper_title, paper_no)
    
    # Check if already exists
    if (PDF_DOWNLOAD_DIR / filename).exists():
        print(f"    ✓ Already exists")
        return True, filename
    
    # Method 1: Direct link from Excel
    if link and not pd.isna(link) and str(link) != 'nan':
        print(f"    → Trying direct link...")
        
        # ArXiv special handling
        arxiv_pdf = try_arxiv_link(link)
        if arxiv_pdf:
            if download_pdf(arxiv_pdf, filename):
                return True, filename
        
        # Try direct link
        if link.lower().endswith('.pdf'):
            if download_pdf(link, filename):
                return True, filename
        
        # Try DOI via Sci-Hub
        doi = extract_doi_from_link(link)
        if doi:
            print(f"    → Trying Sci-Hub with DOI: {doi[:30]}...")
            sci_hub_url = search_sci_hub(paper_title, doi)
            if sci_hub_url:
                if download_pdf(sci_hub_url, filename):
                    return True, filename
    
    # Method 2: Google Scholar search
    print(f"    → Searching Google Scholar...")
    pdf_links = extract_pdf_links_from_google_scholar(paper_title)
    
    for pdf_link in pdf_links[:3]:  # Try first 3 links
        if download_pdf(pdf_link, filename):
            return True, filename
        time.sleep(0.5)  # Small delay between attempts
    
    # Method 3: Sci-Hub with title
    print(f"    → Trying Sci-Hub with title...")
    sci_hub_url = search_sci_hub(paper_title)
    if sci_hub_url:
        if download_pdf(sci_hub_url, filename):
            return True, filename
    
    return False, None

def process_excel_file(excel_file, start_row=0, max_papers=None):
    """Process one Excel file - search and download PDFs."""
    print(f"\n{'='*80}")
    print(f"Processing: {excel_file}")
    print(f"{'='*80}\n")
    
    # Read Excel
    df = pd.read_excel(excel_file)
    
    # Filter rows that need PDFs
    mask = (df['PDF Available'].isna()) | (df['PDF Available'].str.contains('Not Available|Searched', na=False))
    rows_to_process = df[mask]
    
    if start_row > 0:
        rows_to_process = rows_to_process.iloc[start_row:]
    
    if max_papers:
        rows_to_process = rows_to_process.iloc[:max_papers]
    
    print(f"Total papers in file: {len(df)}")
    print(f"Processing: {len(rows_to_process)} papers")
    print(f"Starting from row: {start_row}\n")
    
    downloaded_count = 0
    failed_count = 0
    
    for idx, row in rows_to_process.iterrows():
        paper_title = row['Paper Title']
        link = row.get('Link', None)
        paper_no = row.get('No', idx + 1)
        
        print(f"\n[{paper_no}] {paper_title[:70]}...")
        
        try:
            success, filename = find_and_download_pdf(paper_title, link, paper_no)
            
            if success:
                update_excel_status(excel_file, idx, "Yes")
                downloaded_count += 1
            else:
                update_excel_status(excel_file, idx, "Not Found")
                failed_count += 1
            
            # Rate limiting - be nice to servers
            time.sleep(2)
            
        except Exception as e:
            print(f"    ✗ Error: {e}")
            failed_count += 1
            continue
    
    print(f"\n{'='*80}")
    print(f"Summary for {excel_file}:")
    print(f"  ✓ Downloaded: {downloaded_count}")
    print(f"  ✗ Not found: {failed_count}")
    print(f"  Success rate: {downloaded_count/(downloaded_count+failed_count)*100:.1f}%" if (downloaded_count+failed_count) > 0 else "  N/A")
    print(f"{'='*80}\n")
    
    return downloaded_count, failed_count

def main():
    """Main function to process all Excel files."""
    print("\n" + "="*80)
    print("AUTOMATED PDF DOWNLOADER - Starting...")
    print("="*80)
    
    total_downloaded = 0
    total_failed = 0
    
    # Process first file - limit to 20 papers for testing
    if Path(EXCEL_FILES[0]).exists():
        down, fail = process_excel_file(EXCEL_FILES[0], start_row=0, max_papers=20)
        total_downloaded += down
        total_failed += fail
    
    # Process second file - limit to 20 papers for testing
    if Path(EXCEL_FILES[1]).exists():
        down, fail = process_excel_file(EXCEL_FILES[1], start_row=0, max_papers=20)
        total_downloaded += down
        total_failed += fail
    
    print("\n" + "="*80)
    print("BATCH COMPLETE!")
    print(f"Total Downloaded: {total_downloaded}")
    print(f"Total Failed: {total_failed}")
    print(f"PDFs saved to: {PDF_DOWNLOAD_DIR.absolute()}")
    print("="*80)

if __name__ == "__main__":
    main()

