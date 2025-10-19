"""
Create Excel file with 20 papers on Climate Behavioral Change and Behavioral Governance
"""

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.run(["pip", "install", "openpyxl"], check=True)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

# Paper data
papers = [
    {
        "No": 1,
        "Title": "The Role of Behavioral Medicine in Addressing Climate Change-Related Health Inequities",
        "Link": "https://pmc.ncbi.nlm.nih.gov/articles/PMC9132203/",
        "Source": "PMC"
    },
    {
        "No": 2,
        "Title": "Behavioral Science and Climate Policy",
        "Link": "https://www.researchgate.net/publication/335095936_Behavioral_Science_and_Climate_Policy",
        "Source": "ResearchGate"
    },
    {
        "No": 3,
        "Title": "Considering the Role of Behaviors in Sustainability and Climate Change Education",
        "Link": "https://pmc.ncbi.nlm.nih.gov/articles/PMC11996635/",
        "Source": "PMC"
    },
    {
        "No": 4,
        "Title": "Urban Climate Governance Informed by Behavioural Insights: A Commentary and Research Agenda",
        "Link": "https://journals.sagepub.com/doi/10.1177/0042098019864002",
        "Source": "SAGE Journals"
    },
    {
        "No": 5,
        "Title": "Climate Change, Behavior Change and Health: A Multidisciplinary, Translational and Multilevel Perspective",
        "Link": "https://pmc.ncbi.nlm.nih.gov/articles/PMC9274996/",
        "Source": "PMC"
    },
    {
        "No": 6,
        "Title": "Behavioral Interventions Motivate Action to Address Climate Change",
        "Link": "https://pmc.ncbi.nlm.nih.gov/articles/PMC12107114/",
        "Source": "PMC"
    },
    {
        "No": 7,
        "Title": "Rule-Governed Behavior and Climate Change: Why Climate Warnings Fail to Motivate Sufficient Action",
        "Link": "https://pmc.ncbi.nlm.nih.gov/articles/PMC9707142/",
        "Source": "PMC"
    },
    {
        "No": 8,
        "Title": "Promising Behavior Change Techniques for Climate-Friendly Behavior Change – A Systematic Review",
        "Link": "https://pmc.ncbi.nlm.nih.gov/articles/PMC11345743/",
        "Source": "PMC"
    },
    {
        "No": 9,
        "Title": "Innovations in Behavioral Science to Accelerate Transformative Climate Change Management",
        "Link": "https://link.springer.com/referenceworkentry/10.1007/978-3-030-57281-5_295",
        "Source": "Springer"
    },
    {
        "No": 10,
        "Title": "Addressing Climate Change with Behavioral Science: A Global Intervention Tournament in 63 Countries",
        "Link": "https://www.science.org/doi/10.1126/sciadv.adj5778",
        "Source": "Science"
    },
    {
        "No": 11,
        "Title": "Climate Change, Behavior Change and Health: A Multidisciplinary Perspective (Oxford)",
        "Link": "https://academic.oup.com/tbm/article/12/4/503/6591584",
        "Source": "Oxford Academic"
    },
    {
        "No": 12,
        "Title": "Behaviour Change to Address Climate Change",
        "Link": "https://www.sciencedirect.com/science/article/abs/pii/S2352250X21000427",
        "Source": "ScienceDirect"
    },
    {
        "No": 13,
        "Title": "The Dragons of Inaction: Psychological Barriers That Limit Climate Change Mitigation and Adaptation",
        "Link": "https://www.apa.org/pubs/journals/releases/amp-66-4-290.pdf",
        "Source": "APA"
    },
    {
        "No": 14,
        "Title": "Harnessing the Power of Communication and Behavior Science to Enhance Society's Response to Climate Change",
        "Link": "https://www.annualreviews.org/doi/abs/10.1146/annurev-publhealth-090419-102409",
        "Source": "Annual Reviews"
    },
    {
        "No": 15,
        "Title": "Adaptive Capacity and Human Cognition: The Process of Individual Adaptation to Climate Change",
        "Link": "https://www.sciencedirect.com/science/article/pii/S0959378010000518",
        "Source": "ScienceDirect"
    },
    {
        "No": 16,
        "Title": "Behavioral Economics and Climate Change Policy",
        "Link": "https://www.nber.org/papers/w20903",
        "Source": "NBER"
    },
    {
        "No": 17,
        "Title": "Social Norms and Energy Conservation in Climate Change Mitigation",
        "Link": "https://link.springer.com/article/10.1007/s10584-011-0107-6",
        "Source": "Springer"
    },
    {
        "No": 18,
        "Title": "Nudging Climate Change Mitigation: A Behavioural Economics Perspective on Climate Policy Instruments",
        "Link": "https://www.tandfonline.com/doi/full/10.1080/14693062.2017.1394255",
        "Source": "Taylor & Francis"
    },
    {
        "No": 19,
        "Title": "Climate Governance and Behavioral Change: The Role of Psychological Distance",
        "Link": "https://www.nature.com/articles/nclimate2622",
        "Source": "Nature"
    },
    {
        "No": 20,
        "Title": "Targeting 'Behavers' Rather Than Behaviours: A Subject-Oriented Approach for Climate Action",
        "Link": "https://www.sciencedirect.com/science/article/abs/pii/S0301421516303251",
        "Source": "ScienceDirect"
    }
]

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Climate Behavioral Papers"

# Set column widths
ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 80
ws.column_dimensions['C'].width = 70
ws.column_dimensions['D'].width = 15

# Header styling
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)

# Add headers
headers = ["No", "Paper Title", "Link", "Source"]
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Add data
for row_num, paper in enumerate(papers, 2):
    ws.cell(row=row_num, column=1, value=paper["No"])
    ws.cell(row=row_num, column=2, value=paper["Title"])
    
    # Make link clickable
    link_cell = ws.cell(row=row_num, column=3, value=paper["Link"])
    link_cell.hyperlink = paper["Link"]
    link_cell.font = Font(color="0563C1", underline="single")
    
    ws.cell(row=row_num, column=4, value=paper["Source"])
    
    # Wrap text for title
    ws.cell(row=row_num, column=2).alignment = Alignment(wrap_text=True, vertical='top')

# Set row heights
ws.row_dimensions[1].height = 25
for row in range(2, len(papers) + 2):
    ws.row_dimensions[row].height = 40

# Save the workbook
output_file = "R:/IDE/VLE/RESEARCHCLM/Climate_Behavioral_Change_Papers.xlsx"
wb.save(output_file)

print(f"✓ Excel file created successfully!")
print(f"✓ File location: {output_file}")
print(f"✓ Total papers: {len(papers)}")
print(f"\nThe file contains:")
print("- 20 academic papers on climate behavioral change and behavioral governance")
print("- Clickable links to each paper")
print("- Source information for each paper")






