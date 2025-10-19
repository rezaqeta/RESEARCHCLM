"""
Add case studies from the provided list to Climate_Behavioral_Change_Papers.xlsx
"""

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment
except ImportError:
    import subprocess
    subprocess.run(["pip", "install", "openpyxl"], check=True)
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment

# New case studies extracted from the text
new_case_studies = [
    {
        "title": "Creating the French Behavioral Insights Team - Case - Faculty & Research",
        "link": "https://www.hbs.edu/faculty/Pages/item.aspx?num=55447",
        "organization": "Harvard Business School",
        "pdf": "No PDF"
    },
    {
        "title": "Using Behavioral Insights to Increase Water Revenue in the Recife Metropolitan Region",
        "link": "https://documents.worldbank.org/en/publication/documents-reports/documentdetail/913631620836985463/using-behavioral-insights-to-increase-water-revenue-in-the-recife-metropolitan-region",
        "organization": "World Bank",
        "pdf": "No PDF"
    },
    {
        "title": "Heating reduction as collective action: Impact on attitudes, behavior and energy consumption in a Polish field experiment",
        "link": "https://arxiv.org/abs/2504.11016",
        "organization": "arXiv",
        "pdf": "No PDF"
    },
    {
        "title": "Gap Analysis on Behavioural Research Related to Climate Policy and Interventions",
        "link": "https://www.epa.ie/publications/monitoring--assessment/climate-change/gap-analysis-on-behavioural-research-related-to-climate-policy-and-interventions.php",
        "organization": "EPA Ireland",
        "pdf": "No PDF"
    },
    {
        "title": "A behavioral approach to water conservation: evidence from Costa Rica",
        "link": "https://documents.worldbank.org/en/publication/documents-reports/documentdetail/809801468001190306/a-behavioral-approach-to-water-conservation-evidence-from-costa-rica",
        "organization": "World Bank",
        "pdf": "No PDF"
    },
    {
        "title": "Unlocking the full potential of behavioural insights for policy",
        "link": "https://knowledge4policy.ec.europa.eu/publication/unlocking-full-potential-behavioural-insights-policy_en",
        "organization": "European Commission",
        "pdf": "No PDF"
    },
    {
        "title": "Behavioural Insights Applied to Policy",
        "link": "https://publications.jrc.ec.europa.eu/repository/handle/JRC139824",
        "organization": "EU JRC",
        "pdf": "No PDF"
    },
    {
        "title": "Applying behavioral insights in consumer protection policy",
        "link": "https://documents.worldbank.org/en/publication/documents-reports/documentdetail/634291468331262432/applying-behavioral-insights-in-consumer-protection-policy",
        "organization": "World Bank",
        "pdf": "No PDF"
    },
    {
        "title": "Yes, we can! The effect of collective versus individual action framing on the acceptance of hard climate adaptation policy instruments",
        "link": "https://www.cambridge.org/core/journals/behavioural-public-policy/article/yes-we-can-the-effect-of-collective-versus-individual-action-framing-on-the-acceptance-of-hard-climate-adaptation-policy-instruments/6404F7BFDB8BFF987555D5AE78DA0DBF",
        "organization": "Cambridge Core",
        "pdf": "No PDF"
    },
    {
        "title": "Behavioral Insights for the Water Sector: Improving Outcomes by Changing Mindsets",
        "link": "https://documents.worldbank.org/pt/publication/documents-reports/documentdetail/099192002242351663/p1664190efc6260560b8f104c9edbc7468b",
        "organization": "World Bank",
        "pdf": "No PDF"
    },
    {
        "title": "Behavior Change in Solid Waste Management - A Compendium of Cases",
        "link": "https://documents.worldbank.org/en/publication/documents-reports/documentdetail/099091423124016666/p1773441302811082184c8156db86923f14",
        "organization": "World Bank",
        "pdf": "No PDF"
    },
    {
        "title": "A policy analysis of cleaner technology: a case study of Mondi Limited",
        "link": "https://123dok.com/za/docs/policy-analysis-cleaner-technology-case-study-mondi-limited.10630515",
        "organization": "123dok",
        "pdf": "No PDF"
    },
    {
        "title": "Sustainability as Behavioural Change: Nudging the Good, Discouraging the Bad",
        "link": "https://knowledge.csc.gov.sg/ethos-issue-24/sustainability-as-behavioural-change-nudging-the-good-discouraging-the-bad",
        "organization": "Singapore Civil Service College",
        "pdf": "No PDF"
    },
    {
        "title": "Behavioral Solutions for Successful Subsidy Reform: Research Note",
        "link": "https://documents.worldbank.org/en/publication/documents-reports/documentdetail/099071924103531439/p504124155f2e50ee185dd1fd9a8bfe0f39",
        "organization": "World Bank",
        "pdf": "No PDF"
    },
    {
        "title": "Environmental Institutional Determinants of Climate Behavior in Taiwan's Public Officials",
        "link": "https://www.preprints.org/manuscript/202509.0233/v1",
        "organization": "Preprints.org",
        "pdf": "No PDF"
    },
    {
        "title": "Climate-ADAPT Use Case 20 (Lithuania)",
        "link": "https://climate-adapt.eea.europa.eu/lt/knowledge/climate-adapt-use-cases/use-case-20.pdf",
        "organization": "EEA Climate-ADAPT",
        "pdf": "No PDF"
    },
    {
        "title": "Nudging and habit change for open defecation: new tactics from behavioral science",
        "link": "https://documents.worldbank.org/en/publication/documents-reports/documentdetail/905011467990970572/nudging-and-habit-change-for-open-defecation-new-tactics-from-behavioral-science",
        "organization": "World Bank",
        "pdf": "No PDF"
    },
    {
        "title": "Encouraging Cooperation in Climate Collective Action Problems",
        "link": "https://www.epa.ie/publications/monitoring--assessment/climate-change/encouraging-cooperation-in-climate-collective-action-problems.php",
        "organization": "EPA Ireland",
        "pdf": "No PDF"
    },
    {
        "title": "Impact of public information campaign on citizen behaviors: Vignette experimental study on recycling program in South Korea",
        "link": "https://yonsei.elsevierpure.com/en/publications/impact-of-public-information-campaign-on-citizen-behaviors-vignet",
        "organization": "Yonsei University",
        "pdf": "No PDF"
    },
    {
        "title": "Applying behavioral insights to improve tax collection: experimental evidence from Poland",
        "link": "https://documents.worldbank.org/en/publication/documents-reports/documentdetail/928731497243427428/applying-behavioral-insights-to-improve-tax-collection-experimental-evidence-from-poland",
        "organization": "World Bank",
        "pdf": "No PDF"
    },
    {
        "title": "基于强化理论的公共政策执行优化研究 (Public Policy Implementation Optimization Based on Reinforcement Theory)",
        "link": "https://pdf.hanspub.org/mse_2331036.pdf",
        "organization": "Hans Publishers",
        "pdf": "No PDF"
    },
    {
        "title": "How does central-local policy implementation deviation affect urban residents' green consumption behavior? A multi-agent simulation",
        "link": "https://link.springer.com/article/10.1007/s10668-025-06626-1",
        "organization": "Springer",
        "pdf": "No PDF"
    },
    {
        "title": "Factors Affecting Climate Change Mitigation Policy",
        "link": "https://research.amanote.com/publication/x4tu0nMBKQvf0BhieEQB/factors-affecting-climate-change-mitigation-policy-implementation",
        "organization": "Amanote Research",
        "pdf": "No PDF"
    },
    {
        "title": "The Effective Impact of Behavioral Shifts in Energy, Transport, and Food",
        "link": "https://www.wri.org/research/effective-impact-behavioral-shifts",
        "organization": "World Resources Institute",
        "pdf": "No PDF"
    },
    {
        "title": "Institutional Change Evolution: India's National Action Plan on Climate Change (NAPCC)",
        "link": "https://spp.ucr.edu/document/institutional-change-evolution-indias-national-action-plan-climate-change-napcc",
        "organization": "UC Riverside School of Public Policy",
        "pdf": "No PDF"
    },
    {
        "title": "Should Governments Invest More in Nudging? - Article - Faculty & Research",
        "link": "https://www.hbs.edu/faculty/Pages/item.aspx?num=52754",
        "organization": "Harvard Business School",
        "pdf": "No PDF"
    },
    {
        "title": "Drawing up a national climate change adaptation policy: Feedback from five European case studies (Finnish)",
        "link": "https://climate-adapt.eea.europa.eu/fi/metadata/publications/drawing-up-a-national-climate-change-adaptation-policy-feedback-from-five-european-case-studies",
        "organization": "EEA Climate-ADAPT",
        "pdf": "No PDF"
    },
    {
        "title": "Policy recommendations for community empowerment in climate change adaptation: Go Green Pandawara Group campaign (SDG 13)",
        "link": "https://doaj.org/article/eec2b2ec59f44a13a74fe4b7837793f6",
        "organization": "DOAJ",
        "pdf": "No PDF"
    },
    {
        "title": "Understanding public support for Net Zero policies",
        "link": "https://www.bi.team/publications/understanding-public-support-for-net-zero-policies",
        "organization": "UK Behavioural Insights Team",
        "pdf": "No PDF"
    },
    {
        "title": "Enhancing sustainable consumer behaviour through Nudging: Insights from a field experiment on adoption of electronic receipts",
        "link": "https://www.sciencedirect.com/science/article/pii/S1567422325000730",
        "organization": "ScienceDirect",
        "pdf": "No PDF"
    },
    {
        "title": "Four lessons on the interaction between climate change mitigation policies and social behaviour",
        "link": "https://www.realinstitutoelcano.org/en/analyses/four-lessons-on-the-interaction-between-climate-change-mitigation-policies-and-social-behaviour",
        "organization": "Real Instituto Elcano",
        "pdf": "No PDF"
    },
    {
        "title": "Behaviour Change Strategies and Performance of Corporate Organisations: A Case of Total Kenya Limited",
        "link": "https://erepo.usiu.ac.ke/handle/11732/7034",
        "organization": "USIU-Africa",
        "pdf": "No PDF"
    },
    {
        "title": "Enabling behaviour change: Community-based social marketing strategy in Malmö",
        "link": "https://www.academia.edu/23510984/Enabling_behaviour_change_Community_based_social_marketing_strategy_in_Malm%C3%B6",
        "organization": "Academia.edu",
        "pdf": "No PDF"
    },
    {
        "title": "The Long-Term Impact of Policy Intervention on Intention to Reduce Plastic Bag Usage in Israel",
        "link": "https://www.mdpi.com/2071-1050/17/7/3055",
        "organization": "MDPI Sustainability",
        "pdf": "No PDF"
    },
    {
        "title": "Climate Safe Neighborhoods: A community collaboration for a more climate-resilient future",
        "link": "https://www.communitysci.org/resource/climate-safe-neighborhoods",
        "organization": "Community Science",
        "pdf": "No PDF"
    },
    {
        "title": "Managing Sustainability and Carbon-Neutrality in the Public Administration—Case Report of a German State Institution",
        "link": "https://www.mdpi-res.com/d_attachment/sustainability/sustainability-13-04146/article_deploy/sustainability-13-04146.pdf",
        "organization": "MDPI",
        "pdf": "No PDF"
    },
    {
        "title": "Assessing coherence between sector polices and climate compatible development: opportunities for triple wins",
        "link": "https://www.cccep.ac.uk/publication/assessing-coherence-between-sector-polices-and-climate-compatible-development-opportunities-for-triple-wins",
        "organization": "CCCEP",
        "pdf": "No PDF"
    },
    {
        "title": "Case Study Analysis of Improving Environmental Ethics Using a Collaboration Toolkit",
        "link": "https://link.springer.com/chapter/10.1007/978-3-031-61660-0_15",
        "organization": "Springer",
        "pdf": "No PDF"
    },
    {
        "title": "Climate-Resilient Forest Restoration on Public Lands in Minnesota's Northwoods Region: A Behavior Change Case Study",
        "link": "https://www.scispace.com/papers/climate-resilient-forest-restoration-on-public-lands-in-kpe4i7x7vb",
        "organization": "SciSpace",
        "pdf": "No PDF"
    },
    {
        "title": "London Congestion Charge: Behavioral Traffic Reduction Case Study",
        "link": "https://tfl.gov.uk/corporate/publications-and-reports/congestion-charge",
        "organization": "Transport for London",
        "pdf": "No PDF"
    },
    {
        "title": "Copenhagen Cycling City: Infrastructure and Behavior Change",
        "link": "https://international.kk.dk/artikel/cycling-facts",
        "organization": "City of Copenhagen",
        "pdf": "No PDF"
    },
    {
        "title": "Singapore ERP System: Dynamic Pricing Behavioral Response",
        "link": "https://www.lta.gov.sg/content/ltagov/en/getting_around/driving_in_singapore/erp.html",
        "organization": "Singapore LTA",
        "pdf": "No PDF"
    },
    {
        "title": "Oslo EV Incentives: Policy Impact on Vehicle Choice",
        "link": "https://elbil.no/english/norwegian-ev-policy/",
        "organization": "Norwegian EV Association",
        "pdf": "No PDF"
    },
    {
        "title": "Barcelona Superblocks: Behavioral Street Space Reclaiming",
        "link": "https://ajuntament.barcelona.cat/superilles/en/",
        "organization": "Barcelona City Council",
        "pdf": "No PDF"
    },
    {
        "title": "Climate-ADAPT Use Case 09 (Lithuania)",
        "link": "https://climate-adapt.eea.europa.eu/lt/knowledge/climate-adapt-use-cases/use-case-09.pdf",
        "organization": "EEA Climate-ADAPT",
        "pdf": "No PDF"
    },
    {
        "title": "Behavioural Science for Sustainable Tourism: Insights and policy considerations for greener tourism",
        "link": "https://oecd-opsi.org/publications/behavioural-science-for-sustainable-tourism-insights-and-policy-considerations-for-greener-tourism",
        "organization": "OECD OPSI",
        "pdf": "No PDF"
    },
    {
        "title": "Climate change education through the lens of behavioral economics: A systematic review",
        "link": "https://hal.science/hal-04692537/document",
        "organization": "HAL Science",
        "pdf": "No PDF"
    },
    {
        "title": "The Intention-Behavior Gap in Climate Change Adaptation",
        "link": "https://www.zew.de/en/publications/the-intention-behavior-gap-in-climate-change-adaptation-1",
        "organization": "ZEW Germany",
        "pdf": "No PDF"
    },
    {
        "title": "Exploring Homeowners' Attitudes and Climate-Smart Renovation Decisions: A Case Study in Kronoberg, Sweden",
        "link": "https://www.diva-portal.org/smash/get/diva2:1953452/FULLTEXT01.pdf",
        "organization": "DiVA Portal",
        "pdf": "No PDF"
    },
    {
        "title": "Argentina - Can short term incentives change long term behavior?",
        "link": "https://documents.worldbank.org/en/publication/documents-reports/documentdetail/621961467997641913/argentina-pueden-los-incentivos-de-corto-plazo-cambiar-el-comportamiento-en-el-largo-plazo",
        "organization": "World Bank",
        "pdf": "No PDF"
    },
    {
        "title": "A economia comportamental e a estratégia de nudge da Havan: análise entre os governos Bolsonaro e Lula",
        "link": "https://repositorio.animaeducacao.com.br/items/2a0b6bcc-b7e9-4cb3-9f68-3eff8ce0af90",
        "organization": "Anima Educação Repository",
        "pdf": "No PDF"
    },
    {
        "title": "Gobernanza, cambio climático y soluciones basadas en la naturaleza en Vitoria-Gasteiz",
        "link": "https://recyt.fecyt.es/index.php/CyTET/article/view/107981",
        "organization": "FECYT Spain",
        "pdf": "No PDF"
    },
    {
        "title": "Using Behavioural Insights to Reduce Recreation Impacts on Wildlife: Case Studies from Thames Basin Heaths and Solent",
        "link": "https://publications.naturalengland.org.uk/publication/4742851775954944",
        "organization": "Natural England",
        "pdf": "No PDF"
    },
    {
        "title": "Using Behavioral Insights to Increase Participation in Social Services Programs: A Case Study",
        "link": "https://www.acf.gov/opre/report/using-behavioral-insights-increase-participation-social-services-programs-case-study",
        "organization": "US ACF",
        "pdf": "No PDF"
    },
    {
        "title": "Behavioural insights into energy policy making",
        "link": "https://c2e2.unepccc.org/kms_object/behavioural-insights-into-energy-policy-making",
        "organization": "Copenhagen Centre on Energy Efficiency",
        "pdf": "No PDF"
    },
    {
        "title": "Behavioural Insights (BI) for EU: a case study on embedding BI into public health policy-making",
        "link": "https://publications.jrc.ec.europa.eu/repository/handle/JRC137815",
        "organization": "EU JRC",
        "pdf": "No PDF"
    },
    {
        "title": "Testing the effect of defaults on the thermostat settings of OECD employees",
        "link": "https://econpapers.repec.org/article/eeeeneeco/v_3a39_3ay_3a2013_3ai_3ac_3ap_3a128-134.htm",
        "organization": "Energy Economics Journal",
        "pdf": "No PDF"
    },
    {
        "title": "Creating the French Behavioral Insights Team (HBR Case)",
        "link": "https://store.hbr.org/product/creating-the-french-behavioral-insights-team/919015",
        "organization": "Harvard Business Review",
        "pdf": "No PDF"
    },
    {
        "title": "Behavioural Insights and Monitoring Energy Efficiency",
        "link": "https://www.vringer.nl/docs/2017%2005%2029%20Behavioural%20Insights%20and%20Monitoring%20Energy%20Efficiency%20%28accepted%29.pdf",
        "organization": "Vringer Research",
        "pdf": "No PDF"
    },
    {
        "title": "CHEMIN: Contexte apprenant et incitations de type nudge pour modes de vie durables",
        "link": "https://temis.documentation.developpement-durable.gouv.fr/document.html?id=Temis-0085752",
        "organization": "French Ministry of Environment",
        "pdf": "No PDF"
    },
    {
        "title": "Shaping environmental policy in a citizen-friendly manner",
        "link": "https://www.umweltbundesamt.de/publikationen/shaping-environmental-policy-in-a-citizen-friendly",
        "organization": "German Environment Agency",
        "pdf": "No PDF"
    },
    {
        "title": "Energy Efficiency through Tenant Engagement: A Pilot Behavioral Program for Multifamily Buildings",
        "link": "https://c2e2.unepccc.org/kms_object/energy-efficiency-through-tenant-engagement-a-pilot-behavioral-program-for-multifamily-buildings",
        "organization": "Copenhagen Centre on Energy Efficiency",
        "pdf": "No PDF"
    },
    {
        "title": "Behavior Change Interventions in Practice - Rare",
        "link": "https://behavior.rare.org/behavior-change-in-practice",
        "organization": "Rare Conservation",
        "pdf": "No PDF"
    },
    {
        "title": "Conflict-sensitive adaptation governance: Assessing Kenya's County Climate Change Fund",
        "link": "https://reliefweb.int/report/kenya/conflict-sensitive-adaptation-governance-assessing-kenyas-county-climate-change-fund",
        "organization": "ReliefWeb",
        "pdf": "No PDF"
    },
    {
        "title": "REF Case Study: Behavioral Insights and Climate Policy",
        "link": "https://impact.ref.ac.uk/casestudies/CaseStudy.aspx?Id=38529",
        "organization": "UK REF",
        "pdf": "No PDF"
    },
    {
        "title": "Policy mixes more likely to succeed in fostering sustainable, climate-friendly behaviour",
        "link": "https://joint-research-centre.ec.europa.eu/jrc-news-and-updates/policy-mixes-more-likely-succeed-fostering-sustainable-climate-friendly-behaviour-2024-01-18_en",
        "organization": "EU JRC",
        "pdf": "No PDF"
    },
    {
        "title": "Behavioral Landscape Analysis for Climate Action",
        "link": "https://etcc-ca.com/reports/behavioral-landscape-analysis",
        "organization": "ETCC Canada",
        "pdf": "No PDF"
    },
    {
        "title": "Behavior Change for Nature - Rare Report",
        "link": "https://rare.org/report/behavior-change-for-nature",
        "organization": "Rare Conservation",
        "pdf": "No PDF"
    },
    {
        "title": "Actions to boost Energy Efficiency and Indoor Air Quality: Case studies in Italian schools",
        "link": "https://c2e2.unepccc.org/kms_object/actions-to-boost-energy-efficiency-and-indoor-air-quality-case-studies-in-italian-schools",
        "organization": "Copenhagen Centre on Energy Efficiency",
        "pdf": "No PDF"
    },
    {
        "title": "Using Behavioral Insights to Incentivize Cycling: Results from a Field Experiment",
        "link": "https://www.toi.no/articles-9000/using-behavioral-insights-to-incentivize-cycling-results-from-a-field-experiment-article37081-814.html",
        "organization": "TOI Norway",
        "pdf": "No PDF"
    },
    {
        "title": "Nudging for environmental sustainability: behavioral insights from an on-field experiment #iNUDGEBarletta",
        "link": "https://www.academia.edu/126361006/Nudging_for_environmental_sustainability_behavioral_insights_from_an_on_field_experiment_iNUDGEBarletta",
        "organization": "Academia.edu",
        "pdf": "No PDF"
    },
    {
        "title": "Fostering Green Behavior in the Workplace: The Role of Ethical Climate, Motivation States, and Environmental Knowledge",
        "link": "https://www.mdpi.com/2071-1050/17/9/4083",
        "organization": "MDPI Sustainability",
        "pdf": "No PDF"
    },
    {
        "title": "Creating Pro-Environmental Behavior Change: Economic Incentives or Norm-Nudges?",
        "link": "https://papers.ssrn.com/sol3/papers.cfm?abstract_id=4968244",
        "organization": "SSRN",
        "pdf": "No PDF"
    },
    {
        "title": "Climate change and individual behavior",
        "link": "https://www.bundesbank.de/en/publications/research/discussion-papers/climate-change-and-individual-behavior-844910",
        "organization": "Bundesbank Germany",
        "pdf": "No PDF"
    },
    {
        "title": "Behavioral Facilitation of a Transition to Energy Efficient and Low-Carbon Residential Buildings",
        "link": "https://doaj.org/article/b17c9c9aa3904f049e4b98418df3683d",
        "organization": "DOAJ",
        "pdf": "No PDF"
    },
    {
        "title": "From Eco-Anxiety to Eco-Paralysis: A Case Study on Behavioral Responses to Climate Change in Healthcare Professionals",
        "link": "https://www.em-consulte.com/article/1760744",
        "organization": "EM Consulte",
        "pdf": "No PDF"
    },
    {
        "title": "Explaining Success and Failure in Climate Policies: Developing Theory through German Case Studies",
        "link": "https://academicworks.cuny.edu/hc_pubs/652",
        "organization": "CUNY Academic Works",
        "pdf": "No PDF"
    },
    {
        "title": "Behavior, decisions and ecological transition: experimental approaches with policy implications",
        "link": "https://revistasice.com/index.php/CICE/article/download/7505/7552/8685",
        "organization": "ICE Journal Spain",
        "pdf": "No PDF"
    },
    {
        "title": "A Case Study of British Columbia's Carbon Tax",
        "link": "https://closup.umich.edu/research/student-working-papers/case-study-british-columbias-carbon-tax",
        "organization": "University of Michigan",
        "pdf": "No PDF"
    },
    {
        "title": "Climate policy support as a tool to control others' (but not own) environmental behavior",
        "link": "https://pmc.ncbi.nlm.nih.gov/articles/PMC9216538",
        "organization": "PMC",
        "pdf": "No PDF"
    },
    {
        "title": "Urban Settlement and Behavioural Exchange towards Environmental Policy in Southeast Nigeria",
        "link": "https://ojs.mruni.eu/ojs/public-policy-and-administration/article/view/6944",
        "organization": "Mykolas Romeris University",
        "pdf": "No PDF"
    },
    {
        "title": "Harnessing Behavioral Science to Design Disposable Bag Regulations",
        "link": "https://wagner.nyu.edu/impact/research/publications/harnessing-behavioral-science-design-disposable-bag-regulations-0",
        "organization": "NYU Wagner",
        "pdf": "No PDF"
    },
    {
        "title": "GreenComp in practice: case studies on the use of the European competence framework",
        "link": "https://publications.jrc.ec.europa.eu/repository/handle/JRC140836",
        "organization": "EU JRC",
        "pdf": "No PDF"
    },
    {
        "title": "Transformative Innovation for better Climate Change Adaptation - Case Study: Turku, Southwest Finland",
        "link": "https://place-based-innovation.ec.europa.eu/publications/transformative-innovation-better-climate-change-adaptation-case-study-turku-southwest-finland_en",
        "organization": "EU Place-Based Innovation",
        "pdf": "No PDF"
    },
    {
        "title": "Entrepreneur-led climate adaptation in Kenya",
        "link": "https://www.climatecompatiblegrowth.com/entrepreneur-led-climate-adaptation-in-kenya",
        "organization": "Climate Compatible Growth",
        "pdf": "No PDF"
    },
    {
        "title": "Behavior-Based Energy Efficiency: A Case Study of the Oakland EcoBlock",
        "link": "https://escholarship.org/uc/item/63k9r489",
        "organization": "UC eScholarship",
        "pdf": "No PDF"
    },
    {
        "title": "Transformative Innovation for Climate Change Adaptation – Case study: Espoo, Finland",
        "link": "https://place-based-innovation.ec.europa.eu/publications/transformative-innovation-climate-change-adaptation-case-study-espoo-finland_en",
        "organization": "EU Place-Based Innovation",
        "pdf": "No PDF"
    },
    {
        "title": "Climate Adaptation at the Local Scale: Using Federal Climate Adaptation Policy Regimes to Enhance Climate Services",
        "link": "https://www.mdpi-res.com/d_attachment/sustainability/sustainability-14-08135/article_deploy/sustainability-14-08135.pdf",
        "organization": "MDPI Sustainability",
        "pdf": "No PDF"
    },
    {
        "title": "Scenarios and Sustainability - A Swedish Case Study of Adaptation Tools for Local Decision-Makers",
        "link": "https://www.konj.se/download/18.4ee9b512150ed5e093b9070a/1447246077467/Working-Paper-124-Scenarios-and-Sustainability-A-Swedish-Case-Study-of-Adaptation-Tools-for-Local-Decision-Makers.pdf",
        "organization": "Swedish National Institute of Economic Research",
        "pdf": "No PDF"
    },
    {
        "title": "Responsible Pro-Environmental Management in an Organization: a Case Study",
        "link": "https://sdr.tu.koszalin.pl/info/article/KUT0271cb10bb0e42dfa9836cb58999c431",
        "organization": "Koszalin University",
        "pdf": "No PDF"
    },
    {
        "title": "The Behavioral Effects of Corporate GHG Emissions Disclosures",
        "link": "https://scholarship.law.columbia.edu/sabin_climate_change/256",
        "organization": "Columbia Law School",
        "pdf": "No PDF"
    },
    {
        "title": "'Stop Smoking Like a Chimney!' Rising to the Challenge of Decarbonization - Coopservice Case",
        "link": "https://www.thecasecentre.org/products/view?id=197137",
        "organization": "The Case Centre",
        "pdf": "No PDF"
    },
    {
        "title": "Fiscalité environnementale (Environmental Taxation)",
        "link": "https://sciencespo.hal.science/hal-04615404v1/document",
        "organization": "Sciences Po HAL",
        "pdf": "No PDF"
    },
    {
        "title": "環境感知對環境公共財支付意願及政策偏好的影響 (Environmental Perception Impact on Willingness to Pay)",
        "link": "https://ndltd.ncl.edu.tw/cgi-bin/gs32/gsweb.cgi/login?o=dnclcdr&s=id=\"112NTPU0303005\".&searchmode=basic",
        "organization": "Taiwan Digital Library",
        "pdf": "No PDF"
    },
    {
        "title": "2．研究方法 (Research Methods - Japanese)",
        "link": "https://123dok.com/jp/docs/2%E7%A0%94%E7%A9%B6%E6%96%B9%E6%B3%95.10511251",
        "organization": "123dok Japan",
        "pdf": "No PDF"
    },
    {
        "title": "A Behavioural Multi-criteria Decision Making Framework for Corporate Climate Change Response",
        "link": "https://123dok.com/za/docs/behavioural-criteria-decision-making-framework-corporate-climate-response.10302447",
        "organization": "123dok",
        "pdf": "No PDF"
    },
    {
        "title": "Influência Das Políticas Ambientais No Desempenho (Environmental Policy Influence on Performance)",
        "link": "https://research.amanote.com/publication/qou-0nMBKQvf0BhikOYd/influncia-das-polticas-ambientais-no-desempenho-empresarial-econmico-e-socioambiental",
        "organization": "Amanote Research",
        "pdf": "No PDF"
    }
]

print(f"\n🔍 پردازش {len(new_case_studies)} مورد جدید...")
print(f"   Processing {len(new_case_studies)} new entries...")

# Load existing Excel file
excel_file = "R:/IDE/VLE/RESEARCHCLM/Climate_Behavioral_Change_Papers.xlsx"
wb = load_workbook(excel_file)
ws = wb.active

# Get current last row
last_row = ws.max_row
print(f"\n📊 ردیف‌های موجود: {last_row}")

# Add new case studies
added = 0
for entry in new_case_studies:
    row_num = last_row + added + 1
    
    ws.cell(row=row_num, column=1, value=last_row + added)  # Number
    ws.cell(row=row_num, column=2, value=entry['title'])  # Title
    
    # Link (clickable)
    link_cell = ws.cell(row=row_num, column=3, value=entry['link'])
    link_cell.hyperlink = entry['link']
    link_cell.font = Font(color="0563C1", underline="single")
    
    ws.cell(row=row_num, column=4, value=entry['organization'])  # Organization
    ws.cell(row=row_num, column=5, value="Case Study")  # Type
    ws.cell(row=row_num, column=6, value=entry['pdf'])  # PDF Available
    
    # Format
    ws.cell(row=row_num, column=2).alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[row_num].height = 40
    
    added += 1
    if added % 10 == 0:
        print(f"   ✓ اضافه شد: {added} مورد")

# Save
wb.save(excel_file)

print(f"\n✅ کامل شد! / Complete!")
print(f"\n📊 خلاصه / Summary:")
print(f"   قبل / Before: {last_row} ردیف")
print(f"   اضافه شده / Added: {added} case studies")
print(f"   بعد / After: {last_row + added} ردیف")
print(f"\n📁 فایل به‌روز شده:")
print(f"   {excel_file}")
print(f"\n✅ همه موارد جدید با 'No PDF' علامت‌گذاری شدند")
print(f"✅ All new entries marked with 'No PDF'")





