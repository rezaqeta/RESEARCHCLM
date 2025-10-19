"""
Download PDFs using DOI and Sci-Hub primarily.
"""
import pandas as pd
import requests
from pathlib import Path
import time
import re
from openpyxl import load_workbook
from bs4 import BeautifulSoup
from urllib.parse import quote
import warnings

# Suppress warnings
warnings.filterwarnings('ignore')
requests.packages.urllib3.disable_warnings()

# Configuration
EXCEL_FILES = [
    'source_main_cb.xlsx',
    'Climate_Behavioral_Change_Papers.xlsx'
]
PDF_DOWNLOAD_DIR = Path('sources/metadata/pdfs')
PDF_DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
}

def clean_filename(title, paper_no=None, max_length=80):
    """Create clean filename."""
    filename = re.sub(r'[<>:"/\\|?*]', '', title)
    filename = filename.replace(' ', '_')[:max_length]
    if paper_no:
        filename = f"{paper_no:03d}_{filename}"
    return f"{filename}.pdf"

def extract_doi(text):
    """Extract DOI from text."""
    if not text or pd.isna(text):
        return None
    
    text = str(text)
    
    # Pattern 1: Full DOI URL
    match = re.search(r'doi\.org/(10\.\d{4,}/[^\s\"\'\)]+)', text, re.IGNORECASE)
    if match:
        return match.group(1)
    
    # Pattern 2: DOI prefix
    match = re.search(r'\b(10\.\d{4,}/[^\s\"\'\)]+)', text)
    if match:
        return match.group(1)
    
    return None

def download_from_scihub(doi, filename):
    """Download PDF from Sci-Hub using DOI."""
    scihub_urls = [
        'https://sci-hub.se',
        'https://sci-hub.st', 
        'https://sci-hub.ru',
        'https://sci-hub.mksa.top',
    ]
    
    for base_url in scihub_urls:
        try:
            # Get the Sci-Hub page
            url = f"{base_url}/{doi}"
            print(f"    → Trying: {base_url}")
            
            response = requests.get(url, headers=HEADERS, timeout=15)
            
            if response.status_code != 200:
                continue
            
            # Parse HTML with proper encoding handling
            try:
                soup = BeautifulSoup(response.content, 'lxml')
            except:
                soup = BeautifulSoup(response.text, 'html.parser')
            
            # Find PDF link
            pdf_url = None
            
            # Method 1: iframe with PDF
            iframe = soup.find('iframe', {'id': 'pdf'})
            if iframe and iframe.get('src'):
                pdf_url = iframe['src']
            
            # Method 2: Direct download button
            if not pdf_url:
                for button in soup.find_all('button'):
                    onclick = button.get('onclick', '')
                    if 'location.href' in onclick:
                        match = re.search(r"location\.href='([^']+)'", onclick)
                        if match:
                            pdf_url = match.group(1)
                            break
            
            # Method 3: Direct link
            if not pdf_url:
                link = soup.find('a', href=re.compile(r'\.pdf'))
                if link:
                    pdf_url = link['href']
            
            if not pdf_url:
                continue
            
            # Fix relative URLs
            if pdf_url.startswith('//'):
                pdf_url = 'https:' + pdf_url
            elif pdf_url.startswith('/'):
                pdf_url = base_url + pdf_url
            
            # Download the PDF
            print(f"    → Downloading PDF from: {pdf_url[:60]}...")
            pdf_response = requests.get(pdf_url, headers=HEADERS, timeout=30, stream=True)
            
            if pdf_response.status_code == 200:
                filepath = PDF_DOWNLOAD_DIR / filename
                
                with open(filepath, 'wb') as f:
                    for chunk in pdf_response.iter_content(chunk_size=8192):
                        f.write(chunk)
                
                # Verify it's a PDF
                with open(filepath, 'rb') as f:
                    magic = f.read(4)
                    if magic == b'%PDF':
                        size_kb = filepath.stat().st_size / 1024
                        print(f"    ✓ SUCCESS! Downloaded: {size_kb:.1f} KB")
                        return True
                    else:
                        filepath.unlink()
                        print(f"    ✗ Not a valid PDF")
            
        except Exception as e:
            print(f"    ✗ Error with {base_url}: {str(e)[:40]}")
            continue
    
    return False

def search_crossref_doi(title):
    """Search CrossRef for DOI by title."""
    try:
        url = f"https://api.crossref.org/works?query.title={quote(title)}&rows=1"
        response = requests.get(url, timeout=10)
        
        if response.status_code == 200:
            data = response.json()
            if data['message']['items']:
                doi = data['message']['items'][0].get('DOI')
                if doi:
                    print(f"    → Found DOI via CrossRef: {doi}")
                    return doi
    except:
        pass
    return None

def update_excel(excel_file, row_idx, status):
    """Update Excel status."""
    try:
        wb = load_workbook(excel_file)
        ws = wb.active
        ws.cell(row=row_idx + 2, column=6, value=status)
        wb.save(excel_file)
    except Exception as e:
        print(f"    ⚠ Excel update error: {e}")

def process_paper(paper_title, link, paper_no, excel_file, row_idx):
    """Process a single paper."""
    print(f"\n[{paper_no}] {paper_title[:65]}...")
    
    filename = clean_filename(paper_title, paper_no)
    filepath = PDF_DOWNLOAD_DIR / filename
    
    # Check if already exists
    if filepath.exists():
        print(f"    ✓ Already exists")
        update_excel(excel_file, row_idx, "Yes")
        return True
    
    # Step 1: Extract DOI from link
    doi = extract_doi(link) if link else None
    
    if doi:
        print(f"    → DOI found: {doi}")
        if download_from_scihub(doi, filename):
            update_excel(excel_file, row_idx, "Yes")
            return True
    else:
        print(f"    → No DOI in link")
    
    # Step 2: Search CrossRef for DOI
    print(f"    → Searching CrossRef for DOI...")
    doi = search_crossref_doi(paper_title)
    
    if doi:
        if download_from_scihub(doi, filename):
            update_excel(excel_file, row_idx, "Yes")
            return True
    
    # Step 3: Try direct Sci-Hub search with title
    print(f"    → Trying Sci-Hub with title...")
    if download_from_scihub(paper_title, filename):
        update_excel(excel_file, row_idx, "Yes")
        return True
    
    print(f"    ✗ Could not find PDF")
    update_excel(excel_file, row_idx, "Not Found")
    return False

def process_excel_file(excel_file, max_papers=30):
    """Process Excel file."""
    print(f"\n{'='*80}")
    print(f"Processing: {excel_file}")
    print(f"{'='*80}")
    
    df = pd.read_excel(excel_file)
    
    # Get rows that need PDFs
    mask = (df['PDF Available'].isna()) | (df['PDF Available'].str.contains('Not Available|Searched|Not Found', na=False))
    rows = df[mask].head(max_papers)
    
    print(f"\nTotal papers: {len(df)}")
    print(f"Processing: {len(rows)} papers\n")
    
    success = 0
    failed = 0
    
    for idx, row in rows.iterrows():
        try:
            result = process_paper(
                row['Paper Title'],
                row.get('Link'),
                row.get('No', idx + 1),
                excel_file,
                idx
            )
            
            if result:
                success += 1
            else:
                failed += 1
            
            # Rate limiting
            time.sleep(3)
            
        except KeyboardInterrupt:
            print("\n\n⚠ Interrupted by user")
            break
        except Exception as e:
            print(f"    ✗ Error: {e}")
            failed += 1
    
    print(f"\n{'='*80}")
    print(f"Results for {excel_file}:")
    print(f"  ✓ Downloaded: {success}")
    print(f"  ✗ Failed: {failed}")
    if success + failed > 0:
        print(f"  Success rate: {success/(success+failed)*100:.1f}%")
    print(f"{'='*80}\n")
    
    return success, failed

def main():
    print("\n" + "="*80)
    print("DOI + SCI-HUB PDF DOWNLOADER")
    print("="*80)
    
    total_success = 0
    total_failed = 0
    
    for excel_file in EXCEL_FILES:
        if Path(excel_file).exists():
            s, f = process_excel_file(excel_file, max_papers=30)
            total_success += s
            total_failed += f
    
    print("\n" + "="*80)
    print("FINAL RESULTS")
    print(f"Total Downloaded: {total_success}")
    print(f"Total Failed: {total_failed}")
    print(f"PDFs in: {PDF_DOWNLOAD_DIR.absolute()}")
    print("="*80)

if __name__ == "__main__":
    main()

