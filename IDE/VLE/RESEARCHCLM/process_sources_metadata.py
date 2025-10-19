"""
پردازش فایل‌های sources و افزودن به اکسل‌ها
Process sources files and add to Excel files
"""

import json
import os
from pathlib import Path
try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment
except ImportError:
    import subprocess
    subprocess.run(["pip", "install", "openpyxl"], check=True)
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment

# Paths
metadata_dir = Path("R:/IDE/VLE/RESEARCHCLM/sources/metadata")
json_dir = metadata_dir / "metadatas_jason"
case_studies_excel = Path("R:/IDE/VLE/RESEARCHCLM/Climate_Behavioral_Change_Papers.xlsx")
sources_excel = Path("R:/IDE/VLE/RESEARCHCLM/source_main_cb.xlsx")

# Keywords for identifying case studies
case_study_keywords = [
    "case study", "case-study", "case studies",
    "field experiment", "field trial", "pilot program",
    "implementation", "intervention study",
    "مطالعه موردی", "مورد مطالعاتی"  # Persian
]

# Keywords for excluding non-case studies
exclude_keywords = [
    "review", "meta-analysis", "meta analysis",
    "systematic review", "literature review",
    "framework", "guideline", "handbook",
    "theory", "theoretical"
]

print("🔍 شروع تحلیل فایل‌های sources...")
print("   Starting analysis of sources files...")

# Load existing Excel files
print("\n📂 بارگذاری اکسل‌های موجود...")
wb_cases = load_workbook(case_studies_excel)
ws_cases = wb_cases.active

wb_sources = load_workbook(sources_excel)
ws_sources = wb_sources.active

# Get current last rows
last_row_cases = ws_cases.max_row
last_row_sources = ws_sources.max_row

print(f"   Case Studies Excel: {last_row_cases} ردیف موجود")
print(f"   Sources Excel: {last_row_sources} ردیف موجود")

# Add "PDF Available" column header if not present
if ws_cases.cell(row=1, column=6).value is None:
    ws_cases.cell(row=1, column=6, value="PDF Available")
    ws_cases.cell(row=1, column=6).font = Font(bold=True, color="FFFFFF", size=12)
    from openpyxl.styles import PatternFill
    ws_cases.cell(row=1, column=6).fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws_cases.cell(row=1, column=6).alignment = Alignment(horizontal='center', vertical='center')
    ws_cases.column_dimensions['F'].width = 15

if ws_sources.cell(row=1, column=6).value is None:
    ws_sources.cell(row=1, column=6, value="PDF Available")
    ws_sources.cell(row=1, column=6).font = Font(bold=True, color="FFFFFF", size=12)
    from openpyxl.styles import PatternFill
    ws_sources.cell(row=1, column=6).fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws_sources.cell(row=1, column=6).alignment = Alignment(horizontal='center', vertical='center')
    ws_sources.column_dimensions['F'].width = 15

# Mark existing rows as "Not Available" for PDF column
print("\n📝 به‌روزرسانی ردیف‌های قبلی...")
for row_num in range(2, last_row_cases + 1):
    if ws_cases.cell(row=row_num, column=6).value is None:
        ws_cases.cell(row=row_num, column=6, value="Not Available")

for row_num in range(2, last_row_sources + 1):
    if ws_sources.cell(row=row_num, column=6).value is None:
        ws_sources.cell(row=row_num, column=6, value="Not Available")

# Read all JSON metadata files
print(f"\n🔍 خواندن فایل‌های JSON از {json_dir}...")
json_files = list(json_dir.glob("*.json"))
print(f"   تعداد فایل‌های JSON: {len(json_files)}")

case_studies_to_add = []
sources_to_add = []

print("\n📊 تحلیل فایل‌ها...")
for idx, json_file in enumerate(json_files, 1):
    try:
        with open(json_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        title = data.get('title', '')
        url = data.get('url', '')
        source = data.get('source', '')
        category = data.get('category', '')
        path = data.get('path', '')
        
        # Check if PDF exists
        if path:
            pdf_path = Path("R:/IDE/VLE/RESEARCHCLM/sources/metadata") / path.replace('data\\pdfs\\', 'pdfs\\').replace('\\', '/')
            has_pdf = "Yes" if pdf_path.exists() else "No"
        else:
            has_pdf = "No"
        
        # Determine if it's a case study
        title_lower = title.lower()
        is_case_study = False
        
        # Check for case study keywords
        for keyword in case_study_keywords:
            if keyword in title_lower:
                is_case_study = True
                break
        
        # Check for exclude keywords
        if is_case_study:
            for keyword in exclude_keywords:
                if keyword in title_lower:
                    is_case_study = False
                    break
        
        # Categorize
        entry = {
            'title': title,
            'url': url,
            'source': source,
            'category': category,
            'pdf': has_pdf
        }
        
        if is_case_study:
            case_studies_to_add.append(entry)
        else:
            sources_to_add.append(entry)
        
        if idx % 100 == 0:
            print(f"   پردازش شده: {idx}/{len(json_files)} فایل")
            
    except Exception as e:
        print(f"   خطا در {json_file.name}: {str(e)}")
        continue

print(f"\n✅ تحلیل کامل شد!")
print(f"   📋 Case Studies یافت شده: {len(case_studies_to_add)}")
print(f"   📚 Sources یافت شده: {len(sources_to_add)}")

# Add case studies to Excel
print(f"\n➕ افزودن {len(case_studies_to_add)} Case Study...")
start_num_cases = last_row_cases

for idx, entry in enumerate(case_studies_to_add, 1):
    row_num = last_row_cases + idx
    
    ws_cases.cell(row=row_num, column=1, value=start_num_cases + idx)  # Number
    ws_cases.cell(row=row_num, column=2, value=entry['title'])  # Title
    
    # Link
    link_cell = ws_cases.cell(row=row_num, column=3, value=entry['url'])
    if entry['url']:
        link_cell.hyperlink = entry['url']
        link_cell.font = Font(color="0563C1", underline="single")
    
    ws_cases.cell(row=row_num, column=4, value=entry['source'])  # Organization
    ws_cases.cell(row=row_num, column=5, value=f"Case Study - {entry['category']}")  # Type
    ws_cases.cell(row=row_num, column=6, value=entry['pdf'])  # PDF Available
    
    # Format
    ws_cases.cell(row=row_num, column=2).alignment = Alignment(wrap_text=True, vertical='top')
    ws_cases.row_dimensions[row_num].height = 40
    
    if idx % 20 == 0:
        print(f"   ✓ اضافه شد: {idx} مورد")

# Add sources to Excel
print(f"\n➕ افزودن {len(sources_to_add)} Source...")
start_num_sources = last_row_sources

for idx, entry in enumerate(sources_to_add, 1):
    row_num = last_row_sources + idx
    
    ws_sources.cell(row=row_num, column=1, value=start_num_sources + idx)  # Number
    ws_sources.cell(row=row_num, column=2, value=entry['title'])  # Title
    
    # Link
    link_cell = ws_sources.cell(row=row_num, column=3, value=entry['url'])
    if entry['url']:
        link_cell.hyperlink = entry['url']
        link_cell.font = Font(color="0563C1", underline="single")
    
    ws_sources.cell(row=row_num, column=4, value=entry['source'])  # Organization
    ws_sources.cell(row=row_num, column=5, value=entry['category'])  # Type
    ws_sources.cell(row=row_num, column=6, value=entry['pdf'])  # PDF Available
    
    # Format
    ws_sources.cell(row=row_num, column=2).alignment = Alignment(wrap_text=True, vertical='top')
    ws_sources.row_dimensions[row_num].height = 40
    
    if idx % 50 == 0:
        print(f"   ✓ اضافه شد: {idx} مورد")

# Save files
print(f"\n💾 ذخیره فایل‌ها...")
wb_cases.save(case_studies_excel)
wb_sources.save(sources_excel)

print(f"\n🎉 کامل شد! / Complete!")
print(f"\n📊 خلاصه نهایی / Final Summary:")
print(f"\n📁 Climate_Behavioral_Change_Papers.xlsx")
print(f"   قبل / Before: {last_row_cases} ردیف")
print(f"   اضافه شده / Added: {len(case_studies_to_add)} case studies")
print(f"   بعد / After: {last_row_cases + len(case_studies_to_add)} ردیف")
print(f"   📄 با PDF / With PDF: {sum(1 for e in case_studies_to_add if e['pdf'] == 'Yes')}")
print(f"   ❌ بدون PDF / Without PDF: {sum(1 for e in case_studies_to_add if e['pdf'] == 'No')}")

print(f"\n📁 source_main_cb.xlsx")
print(f"   قبل / Before: {last_row_sources} ردیف")
print(f"   اضافه شده / Added: {len(sources_to_add)} sources")
print(f"   بعد / After: {last_row_sources + len(sources_to_add)} ردیف")
print(f"   📄 با PDF / With PDF: {sum(1 for e in sources_to_add if e['pdf'] == 'Yes')}")
print(f"   ❌ بدون PDF / Without PDF: {sum(1 for e in sources_to_add if e['pdf'] == 'No')}")

print(f"\n✅ همه ردیف‌های قبلی با 'Not Available' علامت‌گذاری شدند")
print(f"✅ All previous rows marked with 'Not Available'")





