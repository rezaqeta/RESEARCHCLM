"""
Use browser automation to download PDFs from the web.
Uses Hyperbrowser MCP tools via Python.
"""
import pandas as pd
from pathlib import Path
import time
from openpyxl import load_workbook
import re

# Configuration
EXCEL_FILES = [
    'source_main_cb.xlsx',
    'Climate_Behavioral_Change_Papers.xlsx'
]
PDF_DOWNLOAD_DIR = Path('sources/metadata/pdfs')
PDF_DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)

def clean_filename(title, paper_no=None, max_length=80):
    """Create clean filename."""
    filename = re.sub(r'[<>:"/\\|?*]', '', title)
    filename = filename.replace(' ', '_')[:max_length]
    if paper_no:
        filename = f"{paper_no:03d}_{filename}"
    return f"{filename}.pdf"

def update_excel(excel_file, row_idx, status):
    """Update Excel status."""
    try:
        wb = load_workbook(excel_file)
        ws = wb.active
        ws.cell(row=row_idx + 2, column=6, value=status)
        wb.save(excel_file)
        return True
    except Exception as e:
        print(f"    ⚠ Excel update error: {e}")
        return False

def generate_search_queries(paper_title):
    """Generate search queries for finding PDFs."""
    queries = [
        f"{paper_title} pdf",
        f"{paper_title} filetype:pdf",
        f'"{paper_title}" pdf download',
    ]
    return queries

def process_excel_file(excel_file, max_papers=50):
    """Process Excel file and prepare download tasks."""
    print(f"\n{'='*80}")
    print(f"Processing: {excel_file}")
    print(f"{'='*80}")
    
    df = pd.read_excel(excel_file)
    
    # Get rows that need PDFs
    mask = (df['PDF Available'].isna()) | (df['PDF Available'].str.contains('Not Available|Searched|Not Found', na=False))
    rows = df[mask].head(max_papers)
    
    print(f"\nTotal papers: {len(df)}")
    print(f"Need to process: {len(rows)} papers\n")
    
    tasks = []
    
    for idx, row in rows.iterrows():
        paper_title = row['Paper Title']
        link = row.get('Link')
        paper_no = row.get('No', idx + 1)
        
        filename = clean_filename(paper_title, paper_no)
        filepath = PDF_DOWNLOAD_DIR / filename
        
        # Skip if already exists
        if filepath.exists():
            print(f"[{paper_no}] ✓ Already exists: {paper_title[:50]}...")
            update_excel(excel_file, idx, "Yes")
            continue
        
        task = {
            'paper_no': paper_no,
            'paper_title': paper_title,
            'link': link,
            'filename': filename,
            'filepath': str(filepath),
            'excel_file': excel_file,
            'row_idx': idx
        }
        tasks.append(task)
        print(f"[{paper_no}] Queued: {paper_title[:50]}...")
    
    return tasks

def main():
    print("\n" + "="*80)
    print("BROWSER-BASED PDF DOWNLOADER")
    print("="*80)
    
    all_tasks = []
    
    for excel_file in EXCEL_FILES:
        if Path(excel_file).exists():
            tasks = process_excel_file(excel_file, max_papers=50)
            all_tasks.extend(tasks)
    
    print(f"\n{'='*80}")
    print(f"Total tasks to download: {len(all_tasks)}")
    print(f"{'='*80}\n")
    
    # Save tasks to file for browser automation
    import json
    with open('download_tasks.json', 'w', encoding='utf-8') as f:
        json.dump(all_tasks, f, indent=2, ensure_ascii=False)
    
    print(f"✓ Tasks saved to: download_tasks.json")
    print(f"\nNext: Use browser automation to download these PDFs")
    
    return all_tasks

if __name__ == "__main__":
    main()

