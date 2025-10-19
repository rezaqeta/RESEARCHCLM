"""
Sequential PDF downloader - searches, finds, lists, and downloads one by one.
"""
import pandas as pd
import requests
from pathlib import Path
import time
import re
from openpyxl import load_workbook
import json
from bs4 import BeautifulSoup
from urllib.parse import urljoin, quote_plus
import warnings

warnings.filterwarnings('ignore')

PDF_DOWNLOAD_DIR = Path('sources/metadata/pdfs')
PDF_DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
}

def find_pdf_links_in_page(url):
    """Search a webpage for PDF links."""
    pdf_links = []
    
    try:
        response = requests.get(url, headers=HEADERS, timeout=15, allow_redirects=True)
        
        if response.status_code == 200:
            soup = BeautifulSoup(response.content, 'html.parser')
            
            # Find all links
            for link in soup.find_all('a', href=True):
                href = link['href']
                
                # Check if it's a PDF link
                if '.pdf' in href.lower() or 'pdf' in link.get_text().lower():
                    # Make absolute URL
                    if href.startswith('http'):
                        pdf_links.append(href)
                    elif href.startswith('/'):
                        base_url = '/'.join(url.split('/')[:3])
                        pdf_links.append(base_url + href)
                    else:
                        pdf_links.append(urljoin(url, href))
            
            # Also check for meta tags
            for meta in soup.find_all('meta', attrs={'name': 'citation_pdf_url'}):
                if meta.get('content'):
                    pdf_links.append(meta['content'])
    
    except Exception as e:
        pass
    
    return list(set(pdf_links))  # Remove duplicates

def search_google_for_pdf(paper_title):
    """Search Google for PDF of the paper."""
    search_query = f"{paper_title} filetype:pdf"
    search_url = f"https://www.google.com/search?q={quote_plus(search_query)}"
    
    pdf_links = []
    
    try:
        response = requests.get(search_url, headers=HEADERS, timeout=15)
        
        if response.status_code == 200:
            soup = BeautifulSoup(response.content, 'html.parser')
            
            # Find all links in search results
            for link in soup.find_all('a', href=True):
                href = link['href']
                
                # Extract actual URL from Google redirect
                if '/url?q=' in href:
                    match = re.search(r'/url\?q=([^&]+)', href)
                    if match:
                        actual_url = match.group(1)
                        if '.pdf' in actual_url:
                            pdf_links.append(actual_url)
    
    except:
        pass
    
    return pdf_links

def convert_source_to_pdf(url):
    """Convert known academic source URLs to PDF links."""
    if not url or pd.isna(url):
        return []
    
    url = str(url).strip()
    pdf_urls = []
    
    # PubMed Central
    if 'pmc.ncbi.nlm.nih.gov' in url or 'ncbi.nlm.nih.gov/pmc' in url:
        pmc_match = re.search(r'PMC\d+', url)
        if pmc_match:
            pmc_id = pmc_match.group(0)
            pdf_urls.append(f"https://www.ncbi.nlm.nih.gov/pmc/articles/{pmc_id}/pdf/")
    
    # arXiv
    if 'arxiv.org' in url:
        arxiv_match = re.search(r'(\d{4}\.\d{4,5})', url)
        if arxiv_match:
            pdf_urls.append(f"https://arxiv.org/pdf/{arxiv_match.group(1)}.pdf")
    
    # bioRxiv / medRxiv
    if 'biorxiv.org' in url or 'medrxiv.org' in url:
        pdf_urls.append(url.rstrip('/') + '.full.pdf')
    
    # Direct PDF
    if url.endswith('.pdf'):
        pdf_urls.append(url)
    
    return pdf_urls

def download_pdf(url, filepath):
    """Download PDF from URL."""
    try:
        print(f"        → Downloading from: {url[:70]}...")
        
        response = requests.get(url, headers=HEADERS, timeout=30, allow_redirects=True, stream=True)
        
        if response.status_code == 200:
            with open(filepath, 'wb') as f:
                for chunk in response.iter_content(chunk_size=8192):
                    if chunk:
                        f.write(chunk)
            
            # Verify PDF
            with open(filepath, 'rb') as f:
                magic = f.read(4)
                if magic == b'%PDF':
                    size_kb = filepath.stat().st_size / 1024
                    print(f"        ✓ SUCCESS! Downloaded {size_kb:.1f} KB")
                    return True
                else:
                    filepath.unlink()
                    print(f"        ✗ Not a valid PDF file")
                    return False
        else:
            print(f"        ✗ HTTP {response.status_code}")
            return False
    
    except Exception as e:
        print(f"        ✗ Error: {str(e)[:50]}")
        return False

def update_excel(excel_file, row_idx, status):
    """Update Excel status."""
    try:
        wb = load_workbook(excel_file)
        ws = wb.active
        ws.cell(row=row_idx + 2, column=6, value=status)
        wb.save(excel_file)
    except:
        pass

def process_paper(task):
    """Process one paper: search, find PDF, list it, download it."""
    paper_no = task['paper_no']
    paper_title = task['paper_title']
    link = task['link']
    filename = task['filename']
    filepath = Path(task['filepath'])
    excel_file = task['excel_file']
    row_idx = task['row_idx']
    
    print(f"\n{'='*80}")
    print(f"[{paper_no}] {paper_title}")
    print(f"{'='*80}")
    
    # Check if already exists
    if filepath.exists():
        print(f"    ✓ PDF already exists: {filename}")
        update_excel(excel_file, row_idx, "Yes")
        return True
    
    # Step 1: Try to convert source link to PDF
    pdf_candidates = []
    
    if link and link != 'nan':
        print(f"    → Original link: {link}")
        
        # Convert known sources
        converted = convert_source_to_pdf(link)
        if converted:
            pdf_candidates.extend(converted)
            print(f"    → Found {len(converted)} direct PDF link(s) from source")
        
        # Search the page for PDF links
        print(f"    → Searching webpage for PDF links...")
        page_pdfs = find_pdf_links_in_page(link)
        if page_pdfs:
            pdf_candidates.extend(page_pdfs)
            print(f"    → Found {len(page_pdfs)} PDF link(s) on page")
    
    # Step 2: Google search for PDF
    print(f"    → Searching Google for PDF...")
    google_pdfs = search_google_for_pdf(paper_title)
    if google_pdfs:
        pdf_candidates.extend(google_pdfs)
        print(f"    → Found {len(google_pdfs)} PDF link(s) from Google")
    
    # Remove duplicates
    pdf_candidates = list(set(pdf_candidates))
    
    if not pdf_candidates:
        print(f"    ✗ No PDF links found")
        update_excel(excel_file, row_idx, "No PDF Found")
        return False
    
    # Step 3: List all PDF links found
    print(f"\n    📄 Found {len(pdf_candidates)} PDF link(s):")
    for i, pdf_url in enumerate(pdf_candidates, 1):
        print(f"       {i}. {pdf_url[:75]}...")
    
    # Step 4: Try downloading from each link
    print(f"\n    💾 Attempting downloads:")
    
    for i, pdf_url in enumerate(pdf_candidates, 1):
        print(f"      [{i}/{len(pdf_candidates)}]")
        
        success = download_pdf(pdf_url, filepath)
        
        if success:
            update_excel(excel_file, row_idx, "Yes")
            return True
        
        time.sleep(1)  # Small delay between attempts
    
    print(f"\n    ✗ All download attempts failed")
    update_excel(excel_file, row_idx, "Download Failed")
    return False

def main():
    print("\n" + "="*80)
    print("SEQUENTIAL PDF DOWNLOADER - One by One")
    print("="*80)
    
    # Load tasks
    with open('download_tasks.json', 'r', encoding='utf-8') as f:
        tasks = json.load(f)
    
    print(f"\nTotal papers to process: {len(tasks)}")
    print(f"PDF directory: {PDF_DOWNLOAD_DIR.absolute()}\n")
    
    success_count = 0
    fail_count = 0
    
    for i, task in enumerate(tasks):
        try:
            success = process_paper(task)
            
            if success:
                success_count += 1
            else:
                fail_count += 1
            
            # Summary every 10 papers
            if (i + 1) % 10 == 0:
                print(f"\n{'='*80}")
                print(f"📊 PROGRESS: {i+1}/{len(tasks)} papers")
                print(f"   ✓ Success: {success_count} | ✗ Failed: {fail_count}")
                print(f"   Success rate: {success_count/(i+1)*100:.1f}%")
                print(f"{'='*80}")
            
            # Rate limiting - be nice to servers
            time.sleep(3)
            
        except KeyboardInterrupt:
            print("\n\n⚠ Stopped by user")
            break
        except Exception as e:
            print(f"\n    ✗ Unexpected error: {e}")
            fail_count += 1
    
    # Final results
    print(f"\n\n{'='*80}")
    print("🎯 FINAL RESULTS")
    print(f"{'='*80}")
    print(f"Total processed: {success_count + fail_count}")
    print(f"✓ Successfully downloaded: {success_count}")
    print(f"✗ Failed: {fail_count}")
    if success_count + fail_count > 0:
        print(f"📈 Success rate: {success_count/(success_count+fail_count)*100:.1f}%")
    print(f"\n📁 PDFs saved to: {PDF_DOWNLOAD_DIR.absolute()}")
    print(f"{'='*80}")

if __name__ == "__main__":
    main()

