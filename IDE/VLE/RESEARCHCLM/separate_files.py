"""
Separate case studies from papers/reports
Keep case studies in Climate_Behavioral_Change_Papers.xlsx
Move papers and reports to source_main_cb.xlsx
"""

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    import subprocess
    subprocess.run(["pip", "install", "openpyxl"], check=True)
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

# Load the original file
original_file = "R:/IDE/VLE/RESEARCHCLM/Climate_Behavioral_Change_Papers.xlsx"
wb_original = load_workbook(original_file)
ws_original = wb_original.active

# Create new workbook for papers and reports
wb_new = Workbook()
ws_new = wb_new.active
ws_new.title = "Papers and Reports"

# Copy header row to new file
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)

for col in range(1, 6):
    cell_original = ws_original.cell(row=1, column=col)
    cell_new = ws_new.cell(row=1, column=col)
    cell_new.value = cell_original.value
    cell_new.font = header_font
    cell_new.fill = header_fill
    cell_new.alignment = Alignment(horizontal='center', vertical='center')

# Set column widths for new file
ws_new.column_dimensions['A'].width = 5
ws_new.column_dimensions['B'].width = 80
ws_new.column_dimensions['C'].width = 70
ws_new.column_dimensions['D'].width = 15
ws_new.column_dimensions['E'].width = 18

# Lists to track entries
case_studies = []
papers_reports = []

print("\n🔍 Analyzing entries...")

# Read all entries from original file (skip header)
for row_num in range(2, ws_original.max_row + 1):
    entry_type = ws_original.cell(row=row_num, column=5).value  # Column E is Type
    
    entry = {
        'row': row_num,
        'number': ws_original.cell(row=row_num, column=1).value,
        'title': ws_original.cell(row=row_num, column=2).value,
        'link': ws_original.cell(row=row_num, column=3).value,
        'organization': ws_original.cell(row=row_num, column=4).value,
        'type': entry_type
    }
    
    # Check if it's a case study
    if entry_type and "Case Study" in str(entry_type):
        case_studies.append(entry)
    else:
        papers_reports.append(entry)

print(f"   ✓ Found {len(case_studies)} case studies")
print(f"   ✓ Found {len(papers_reports)} papers/reports")

# Add papers and reports to new file
print(f"\n📋 Creating source_main_cb.xlsx...")
for idx, entry in enumerate(papers_reports, 1):
    row_num = idx + 1
    
    ws_new.cell(row=row_num, column=1, value=idx)
    ws_new.cell(row=row_num, column=2, value=entry['title'])
    
    # Make link clickable
    link_cell = ws_new.cell(row=row_num, column=3, value=entry['link'])
    if entry['link']:
        link_cell.hyperlink = entry['link']
        link_cell.font = Font(color="0563C1", underline="single")
    
    ws_new.cell(row=row_num, column=4, value=entry['organization'])
    ws_new.cell(row=row_num, column=5, value=entry['type'])
    
    # Wrap text
    ws_new.cell(row=row_num, column=2).alignment = Alignment(wrap_text=True, vertical='top')
    ws_new.row_dimensions[row_num].height = 40

# Save new file with papers and reports
new_file = "R:/IDE/VLE/RESEARCHCLM/source_main_cb.xlsx"
wb_new.save(new_file)
print(f"   ✓ Saved papers and reports to: {new_file}")

# Now recreate original file with only case studies
print(f"\n📊 Updating Climate_Behavioral_Change_Papers.xlsx...")

# Clear all data rows (keep header)
for row_num in range(ws_original.max_row, 1, -1):
    if row_num > 1:  # Don't delete header
        ws_original.delete_rows(row_num)

# Add only case studies back
for idx, entry in enumerate(case_studies, 1):
    row_num = idx + 1
    
    ws_original.cell(row=row_num, column=1, value=idx)
    ws_original.cell(row=row_num, column=2, value=entry['title'])
    
    # Make link clickable
    link_cell = ws_original.cell(row=row_num, column=3, value=entry['link'])
    if entry['link']:
        link_cell.hyperlink = entry['link']
        link_cell.font = Font(color="0563C1", underline="single")
    
    ws_original.cell(row=row_num, column=4, value=entry['organization'])
    ws_original.cell(row=row_num, column=5, value=entry['type'])
    
    # Wrap text
    ws_original.cell(row=row_num, column=2).alignment = Alignment(wrap_text=True, vertical='top')
    ws_original.row_dimensions[row_num].height = 40

# Save updated original file
wb_original.save(original_file)
print(f"   ✓ Updated with case studies only")

# Summary
print(f"\n✅ SUCCESS! Files separated:")
print(f"\n📁 File 1: Climate_Behavioral_Change_Papers.xlsx")
print(f"   📊 Contains: {len(case_studies)} case studies ONLY")
print(f"   🎯 Focus: Real-world behavioral interventions")
print(f"\n📁 File 2: source_main_cb.xlsx")
print(f"   📚 Contains: {len(papers_reports)} papers and reports")
print(f"   📖 Breakdown:")

# Count by type
type_counts = {}
for entry in papers_reports:
    entry_type = entry['type'] or "Unknown"
    type_counts[entry_type] = type_counts.get(entry_type, 0) + 1

for type_name, count in sorted(type_counts.items()):
    print(f"      • {type_name}: {count}")

print(f"\n🎉 Organization complete!")
print(f"   Total entries processed: {len(case_studies) + len(papers_reports)}")





