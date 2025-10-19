"""
Smart PDF downloader that handles common academic sources directly.
Works with PMC, ResearchGate, arXiv, DOI links, etc.
"""
import pandas as pd
import requests
from pathlib import Path
import time
import re
from openpyxl import load_workbook
import json
import warnings

warnings.filterwarnings('ignore')

# Configuration
PDF_DOWNLOAD_DIR = Path('sources/metadata/pdfs')
PDF_DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
}

def convert_to_pdf_link(url):
    """Convert various academic URLs to direct PDF links."""
    if not url or pd.isna(url):
        return None
    
    url = str(url).strip()
    
    # PubMed Central (PMC)
    if 'pmc.ncbi.nlm.nih.gov' in url or 'ncbi.nlm.nih.gov/pmc' in url:
        if '/pdf' not in url:
            # Extract PMC ID
            pmc_match = re.search(r'PMC\d+', url)
            if pmc_match:
                pmc_id = pmc_match.group(0)
                return f"https://www.ncbi.nlm.nih.gov/pmc/articles/{pmc_id}/pdf/"
            else:
                # Try adding /pdf to the end
                return url.rstrip('/') + '/pdf/'
        return url
    
    # arXiv
    if 'arxiv.org' in url:
        if '/pdf/' not in url:
            # Extract arXiv ID
            arxiv_match = re.search(r'(\d{4}\.\d{4,5})', url)
            if arxiv_match:
                return f"https://arxiv.org/pdf/{arxiv_match.group(1)}.pdf"
        return url
    
    # bioRxiv / medRxiv
    if 'biorxiv.org' in url or 'medrxiv.org' in url:
        if '.pdf' not in url:
            return url.rstrip('/') + '.full.pdf'
        return url
    
    # ResearchGate - try to get PDF
    if 'researchgate.net' in url:
        # ResearchGate needs special handling, return as-is for now
        return url
    
    # DOI links
    if 'doi.org' in url:
        doi = re.search(r'10\.\d{4,}/[^\s]+', url)
        if doi:
            # Try Unpaywall API
            return f"https://api.unpaywall.org/v2/{doi.group(0)}?email=research@example.com"
    
    # Already a PDF
    if url.endswith('.pdf'):
        return url
    
    return url

def download_pdf(url, filepath, max_retries=2):
    """Download PDF from URL."""
    for attempt in range(max_retries):
        try:
            response = requests.get(url, headers=HEADERS, timeout=30, allow_redirects=True, stream=True)
            
            if response.status_code == 200:
                # Check content type
                content_type = response.headers.get('Content-Type', '').lower()
                
                # Save file
                with open(filepath, 'wb') as f:
                    for chunk in response.iter_content(chunk_size=8192):
                        if chunk:
                            f.write(chunk)
                
                # Verify it's a PDF
                with open(filepath, 'rb') as f:
                    magic = f.read(4)
                    if magic == b'%PDF':
                        size_kb = filepath.stat().st_size / 1024
                        return True, f"Downloaded ({size_kb:.1f} KB)"
                    else:
                        filepath.unlink()
                        return False, "Not a valid PDF"
            else:
                return False, f"HTTP {response.status_code}"
                
        except Exception as e:
            if attempt == max_retries - 1:
                return False, f"Error: {str(e)[:40]}"
            time.sleep(1)
    
    return False, "Failed after retries"

def try_unpaywall(doi, filepath):
    """Try to get PDF from Unpaywall API."""
    try:
        url = f"https://api.unpaywall.org/v2/{doi}?email=research@example.com"
        response = requests.get(url, timeout=10)
        
        if response.status_code == 200:
            data = response.json()
            
            # Check for open access PDF
            if data.get('is_oa'):
                best_oa = data.get('best_oa_location')
                if best_oa and best_oa.get('url_for_pdf'):
                    pdf_url = best_oa['url_for_pdf']
                    return download_pdf(pdf_url, filepath)
    except:
        pass
    
    return False, "No OA version found"

def update_excel(excel_file, row_idx, status):
    """Update Excel status."""
    try:
        wb = load_workbook(excel_file)
        ws = wb.active
        ws.cell(row=row_idx + 2, column=6, value=status)
        wb.save(excel_file)
    except Exception as e:
        print(f"      ⚠ Excel update failed: {e}")

def process_task(task):
    """Process a single download task."""
    paper_no = task['paper_no']
    paper_title = task['paper_title']
    link = task['link']
    filename = task['filename']
    filepath = Path(task['filepath'])
    excel_file = task['excel_file']
    row_idx = task['row_idx']
    
    print(f"\n[{paper_no}] {paper_title[:60]}...")
    
    # Check if already exists
    if filepath.exists():
        print(f"    ✓ Already exists")
        update_excel(excel_file, row_idx, "Yes")
        return True
    
    # Try direct link
    if link and link != 'nan':
        pdf_url = convert_to_pdf_link(link)
        
        if pdf_url:
            print(f"    → Trying: {pdf_url[:70]}...")
            success, message = download_pdf(pdf_url, filepath)
            
            if success:
                print(f"    ✓ {message}")
                update_excel(excel_file, row_idx, "Yes")
                return True
            else:
                print(f"    ✗ {message}")
        
        # Try Unpaywall if DOI exists
        doi = re.search(r'10\.\d{4,}/[^\s]+', str(link))
        if doi:
            print(f"    → Trying Unpaywall with DOI: {doi.group(0)[:40]}...")
            success, message = try_unpaywall(doi.group(0), filepath)
            
            if success:
                print(f"    ✓ {message}")
                update_excel(excel_file, row_idx, "Yes")
                return True
            else:
                print(f"    ✗ {message}")
    
    print(f"    ✗ Could not download")
    update_excel(excel_file, row_idx, "Download Failed")
    return False

def main():
    print("\n" + "="*80)
    print("SMART PDF DOWNLOADER")
    print("="*80)
    
    # Load tasks
    with open('download_tasks.json', 'r', encoding='utf-8') as f:
        tasks = json.load(f)
    
    print(f"\nTotal tasks: {len(tasks)}")
    print(f"Starting download...\n")
    
    success_count = 0
    fail_count = 0
    
    for i, task in enumerate(tasks):
        try:
            success = process_task(task)
            
            if success:
                success_count += 1
            else:
                fail_count += 1
            
            # Rate limiting
            time.sleep(2)
            
            # Progress update every 10 papers
            if (i + 1) % 10 == 0:
                print(f"\n{'='*80}")
                print(f"Progress: {i+1}/{len(tasks)} papers processed")
                print(f"Success: {success_count} | Failed: {fail_count}")
                print(f"{'='*80}")
            
        except KeyboardInterrupt:
            print("\n\n⚠ Interrupted by user")
            break
        except Exception as e:
            print(f"    ✗ Error: {e}")
            fail_count += 1
    
    # Final summary
    print(f"\n{'='*80}")
    print("FINAL RESULTS")
    print(f"{'='*80}")
    print(f"Total processed: {success_count + fail_count}")
    print(f"✓ Downloaded: {success_count}")
    print(f"✗ Failed: {fail_count}")
    if success_count + fail_count > 0:
        print(f"Success rate: {success_count/(success_count+fail_count)*100:.1f}%")
    print(f"\nPDFs saved to: {PDF_DOWNLOAD_DIR.absolute()}")
    print(f"{'='*80}")

if __name__ == "__main__":
    main()

